from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.http import HttpResponse
from django.apps import apps
import openpyxl

@staff_member_required
def export_report(request):
    if request.method == 'POST':
        selected_tables = request.POST.getlist('tables')
        selected_fields = request.POST.getlist('fields')
        
        if not selected_tables:
            return HttpResponse("Выберите хотя бы одну таблицу", status=400)
        if not selected_fields:
            return HttpResponse("Выберите хотя бы одно поле", status=400)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        row = 1
        for table_key in selected_tables:
            try:
                model = apps.get_model(table_key)
            except LookupError:
                continue
            model_fields = [f for f in selected_fields if f.startswith(table_key + '.')]
            if not model_fields:
                continue
            # Заголовки
            col = 1
            for mf in model_fields:
                field_name = mf.split('.')[-1]
                try:
                    verbose = model._meta.get_field(field_name).verbose_name
                except:
                    verbose = field_name
                ws.cell(row=row, column=col, value=f"{model._meta.verbose_name} – {verbose}")
                col += 1
            row += 1
            # Данные
            for obj in model.objects.all().order_by('id'):
                col = 1
                for mf in model_fields:
                    field_name = mf.split('.')[-1]
                    value = getattr(obj, field_name, '')
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    ws.cell(row=row, column=col, value=str(value) if value is not None else '')
                    col += 1
                row += 1
            row += 1  # пустая строка между таблицами
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=report.xlsx'
        wb.save(response)
        return response
    
    # GET: показываем форму (как раньше)
    all_models = apps.get_models()
    models_data = []
    for model in all_models:
        if model._meta.app_label in ('auth', 'sessions', 'admin', 'contenttypes'):
            continue
        models_data.append({
            'key': f"{model._meta.app_label}.{model.__name__}",
            'verbose': model._meta.verbose_name,
            'fields': []
        })
        for field in model._meta.get_fields():
            if field.name in ('id', 'created_at', 'updated_at') or field.is_relation:
                continue
            models_data[-1]['fields'].append({
                'name': field.name,
                'verbose': field.verbose_name
            })
    return render(request, 'reports/export_form.html', {'models': models_data})